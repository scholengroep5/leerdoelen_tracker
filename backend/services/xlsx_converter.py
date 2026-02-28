"""
XLSX naar JSON converter voor GO! Basisonderwijs doelensets.

Converteert een Excel bestand naar het interne JSON formaat van de
Leerdoelen Tracker. Ondersteunt twee bestandsnaamconventies van GO!:

  Oud formaat:  Doelenset_BaO_<vak>.xlsx
  Nieuw formaat: <nr>__<Vak>.xlsx  (bv. 10__ICT.xlsx, 03__Nederlands.xlsx)

De vak ID wordt afgeleid van de sheetnaam (meest betrouwbaar) of
de bestandsnaam als fallback.

Versie-informatie: het GO! bestand bevat geen expliciete versie of datum
in de inhoud. De 'gewijzigd'-timestamp uit de Excel metadata (wb.properties.modified)
is de meest betrouwbare indicator voor updates. Deze wordt opgeslagen in het
JSON als 'bronDatum' zodat de beheerder kan zien wanneer het bestand voor het
laatste door GO! werd bijgewerkt.

Updates ophalen: https://pro.g-o.be/themas/leerplanning/basisonderwijs/nieuw-leerplan-basisonderwijs/
Er is geen publieke API — download de xlsx bestanden manueel en upload ze hier.
"""

import re
import io
import logging

logger = logging.getLogger(__name__)

# Leeftijdskolommen — meerdere varianten want GO! is niet consistent
# '2,5-4' en '3-4' zijn beide gezien voor de jongste groep
LEEFTIJD_KOLOMMEN = ['2,5-4', '3-4', '4-5', '5-6', '6-7', '7-8', '8-9', '9-10', '10-11', '11-12']
# Genormaliseerde naam voor '2,5-4' in de output (consistent met rest van de app)
LEEFTIJD_NORMALISATIE = {'2,5-4': '2,5-4'}  # bewaren zoals het is, frontend toont het zo

# Hiërarchische structuurtypes (van hoog naar laag niveau)
HIERARCHIE_TYPES = ['onderwerp', 'rubriek', 'subrubriek', 'subthema']

# Types die gekoppeld worden aan de vorige doelzin
DOELZIN_KINDEREN = {
    'MIA - titel', 'MIA - aanklikbaar', 'MIA - niet aanklikbaar',
    'te hanteren begrippen', 'voorbeelden - titel', 'voorbeelden - bullet', 'asterisk',
}

# Vak naam mapping — dekt beide bestandsnaamconventies + sheetnamen
VAK_NAMEN = {
    # Op vak ID (intern formaat)
    'doelenset-bao-aardrijkskunde':         'Aardrijkskunde',
    'doelenset-bao-burgerschap':            'Burgerschap',
    'doelenset-bao-frans':                  'Frans',
    'doelenset-bao-geschiedenis':           'Geschiedenis',
    'doelenset-bao-ict':                    'ICT',
    'doelenset-bao-leren-leren':            'Leren leren',
    'doelenset-bao-lichamelijke-opvoeding': 'Lichamelijke opvoeding',
    'doelenset-bao-muzische-vorming':       'Muzische vorming',
    'doelenset-bao-nederlands':             'Nederlands',
    'doelenset-bao-sociale-vaardigheden':   'Sociale vaardigheden',
    'doelenset-bao-wetenschap-techniek':    'Wetenschap en techniek',
    'doelenset-bao-wiskunde':               'Wiskunde',
    # Op sheetnaam (lowercase) — nieuw GO! formaat
    'ict':                    'ICT',
    'nederlands':             'Nederlands',
    'wiskunde':               'Wiskunde',
    'aardrijkskunde':         'Aardrijkskunde',
    'burgerschap':            'Burgerschap',
    'frans':                  'Frans',
    'geschiedenis':           'Geschiedenis',
    'leren leren':            'Leren leren',
    'lichamelijke opvoeding': 'Lichamelijke opvoeding',
    'muzische vorming':       'Muzische vorming',
    'sociale vaardigheden':   'Sociale vaardigheden',
    'wetenschap en techniek': 'Wetenschap en techniek',
    'wetenschap & techniek':  'Wetenschap en techniek',
}

# Mapping van sheetnaam → intern vak ID
SHEET_NAAR_VAK_ID = {
    'ict':                    'doelenset-bao-ict',
    'nederlands':             'doelenset-bao-nederlands',
    'wiskunde':               'doelenset-bao-wiskunde',
    'aardrijkskunde':         'doelenset-bao-aardrijkskunde',
    'burgerschap':            'doelenset-bao-burgerschap',
    'frans':                  'doelenset-bao-frans',
    'geschiedenis':           'doelenset-bao-geschiedenis',
    'leren leren':            'doelenset-bao-leren-leren',
    'lichamelijke opvoeding': 'doelenset-bao-lichamelijke-opvoeding',
    'muzische vorming':       'doelenset-bao-muzische-vorming',
    'sociale vaardigheden':   'doelenset-bao-sociale-vaardigheden',
    'wetenschap en techniek': 'doelenset-bao-wetenschap-techniek',
    'wetenschap & techniek':  'doelenset-bao-wetenschap-techniek',
}


def vak_id_van_sheetnaam(sheetnaam: str) -> str | None:
    """Leidt vak ID af van de sheetnaam (meest betrouwbaar)."""
    key = sheetnaam.strip().lower()
    return SHEET_NAAR_VAK_ID.get(key)


def vak_id_van_bestandsnaam(bestandsnaam: str) -> str:
    """
    Leidt vak ID af van de bestandsnaam als fallback.

    Ondersteunt:
      Oud: Doelenset_BaO_wiskunde.xlsx       → doelenset-bao-wiskunde
      Nieuw: 10__ICT.xlsx                    → doelenset-bao-ict
             03__Nederlands.xlsx             → doelenset-bao-nederlands
             07__Muzische_vorming.xlsx       → doelenset-bao-muzische-vorming
    """
    naam = bestandsnaam
    if naam.lower().endswith('.xlsx'):
        naam = naam[:-5]

    # Verwijder kopie-suffixen
    naam = re.sub(r'[\s_]*-[\s_]*[Cc]opy.*$', '', naam)
    naam = re.sub(r'\s*\(\d+\)\s*$', '', naam)

    # Nieuw formaat: verwijder nummer-prefix (10__, 03__, etc.)
    naam = re.sub(r'^\d+__', '', naam)

    # Vervang underscores en spaties door streepjes, lowercase
    naam = naam.replace('_', '-').replace(' ', '-').lower()

    # Specifieke correcties
    naam = naam.replace('-en-techniek', '-techniek')

    # Opruimen
    naam = re.sub(r'-+', '-', naam).strip('-')

    # Probeer te mappen naar bekend vak ID via sheetnaam-logica
    known = SHEET_NAAR_VAK_ID.get(naam.replace('-', ' '))
    if known:
        return known

    # Fallback: voeg prefix toe als die er nog niet is
    if not naam.startswith('doelenset-bao-'):
        naam = f'doelenset-bao-{naam}'

    return naam


def converteer_xlsx_naar_json(bestandsnaam: str, bestand_inhoud: bytes) -> dict:
    """
    Converteert een Excel bestand (als bytes) naar het interne JSON formaat.

    Args:
        bestandsnaam:   Originele bestandsnaam, bv. "10__ICT.xlsx"
        bestand_inhoud: Raw bytes van het xlsx bestand

    Returns:
        Dict met het volledige vak JSON object, klaar om op te slaan.

    Raises:
        ValueError: Als het bestand niet gelezen kan worden of structuur klopt niet.
        ImportError: Als pandas/openpyxl niet geïnstalleerd zijn.
    """
    try:
        import pandas as pd
        import openpyxl
    except ImportError:
        raise ImportError(
            "pandas en openpyxl zijn niet geïnstalleerd. "
            "Voeg 'pandas' en 'openpyxl' toe aan requirements.txt"
        )

    buf = io.BytesIO(bestand_inhoud)

    # Lees metadata voor versiedatum (beste versie-indicator beschikbaar)
    bron_datum = None
    sheet_naam = None
    try:
        wb = openpyxl.load_workbook(buf, read_only=True)
        sheet_naam = wb.sheetnames[0] if wb.sheetnames else None
        if wb.properties.modified:
            bron_datum = wb.properties.modified.strftime('%Y-%m-%d')
        wb.close()
    except Exception as e:
        logger.warning(f"Kon workbook metadata niet lezen: {e}")

    # Bepaal vak ID: sheetnaam heeft prioriteit over bestandsnaam
    vak_id = None
    if sheet_naam:
        vak_id = vak_id_van_sheetnaam(sheet_naam)
        if vak_id:
            logger.info(f"Vak ID bepaald via sheetnaam '{sheet_naam}': {vak_id}")
    if not vak_id:
        vak_id = vak_id_van_bestandsnaam(bestandsnaam)
        logger.info(f"Vak ID bepaald via bestandsnaam '{bestandsnaam}': {vak_id}")

    # Lees Excel data
    buf.seek(0)
    try:
        df = pd.read_excel(buf, engine='openpyxl')
    except Exception as e:
        raise ValueError(f"Kon Excel bestand niet lezen: {e}")

    # Minimale kolomcheck
    verplichte_kolommen = {'TYPE', 'GO! NR.', 'INHOUD'}
    ontbrekend = verplichte_kolommen - set(df.columns)
    if ontbrekend:
        raise ValueError(
            f"Verplichte kolommen ontbreken: {', '.join(sorted(ontbrekend))}. "
            f"Is dit een geldig GO! doelenset bestand?"
        )

    # KENNISVERWERKING: sommige bestanden (bv. Wiskunde) missen de kolomkop waardoor
    # pandas de kolom 'Unnamed: 12' noemt. Detecteer op basis van inhoud als fallback.
    kennis_kolom = 'KENNISVERWERKING'
    if kennis_kolom not in df.columns:
        ebg_waarden = {'engageren', 'begrijpen', 'gebruiken'}
        for col in df.columns:
            uniek = set(str(v).lower() for v in df[col].dropna().unique())
            if uniek and uniek.issubset(ebg_waarden):
                kennis_kolom = col
                logger.warning(
                    f"KENNISVERWERKING kolom niet gevonden bij naam in '{bestandsnaam}' — "
                    f"gebruik '{col}' op basis van inhoud (engageren/begrijpen/gebruiken)"
                )
                break

    # Bepaal aanwezige leeftijdskolommen (volgorde bewaren)
    aanwezige_leeftijden = [l for l in LEEFTIJD_KOLOMMEN if l in df.columns]

    # Hiërarchie-tracker
    huidige_parents = {niveau: None for niveau in HIERARCHIE_TYPES}
    huidige_doelzin_id = None
    rijen = []
    aantal_doelzinnen = 0

    for idx, row in df.iterrows():
        rij_id = idx + 1

        row_type = row.get('TYPE', '')
        if pd.isna(row_type):
            row_type = ''
        row_type = str(row_type).strip()

        # Leeftijden: True of 1 in de kolom
        leeftijden = [
            lft for lft in aanwezige_leeftijden
            if row.get(lft) is True or row.get(lft) == 1
        ]

        def cel(kolom):
            val = row.get(kolom)
            try:
                if pd.isna(val):
                    return None
            except (TypeError, ValueError):
                pass
            s = str(val).strip() if val is not None else None
            return s if s else None

        rij = {
            'id':               rij_id,
            'type':             row_type or None,
            'goNr':             cel('GO! NR.'),
            'inhoud':           cel('INHOUD'),
            'koNummer':         cel('KO nummer'),
            'koMinimumdoel':    cel('KO minimumdoel'),
            'koOmschrijving':   cel('KO omschrijving'),
            'l4Nummer':         cel('L4 nummer'),
            'l4Minimumdoel':    cel('L4 minimumdoel'),
            'l4Omschrijving':   cel('L4 omschrijving'),
            'l6Nummer':         cel('L6 nummer'),
            'l6Minimumdoel':    cel('L6 minimumdoel'),
            'l6Omschrijving':   cel('L6 omschrijving'),
            'kennisverwerking': cel(kennis_kolom),
            'leeftijden':       leeftijden,
            'vlaggen':          cel('VLAGGEN'),
            'schakels':         cel('SCHAKELS'),
        }

        # Hiërarchie bijhouden
        if row_type in HIERARCHIE_TYPES:
            huidige_parents[row_type] = rij_id
            reset = False
            for niveau in HIERARCHIE_TYPES:
                if reset:
                    huidige_parents[niveau] = None
                if niveau == row_type:
                    reset = True

        # minimumdoel: behandelen als structuurelement (onder subthema, boven doelzin)
        # Bevat nuttige L4/L6 info maar geen eigen GO! nummer — koppelen aan subthema
        if row_type == 'minimumdoel':
            rij['parentId'] = huidige_parents.get('subthema') or huidige_parents.get('onderwerp')

        # Parent koppelen voor doelzinnen
        elif row_type == 'doelzin':
            aantal_doelzinnen += 1
            huidige_doelzin_id = rij_id
            for niveau in reversed(HIERARCHIE_TYPES):
                if huidige_parents[niveau]:
                    rij['parentId'] = huidige_parents[niveau]
                    break

        # MIA en gerelateerde items koppelen aan laatste doelzin
        elif row_type and (row_type.startswith('MIA') or row_type in DOELZIN_KINDEREN):
            rij['parentDoelzinId'] = huidige_doelzin_id

        rijen.append(rij)

    if aantal_doelzinnen == 0:
        raise ValueError(
            "Geen doelzinnen gevonden (type='doelzin'). "
            "Controleer of dit het juiste bestand is."
        )

    vak_naam = VAK_NAMEN.get(vak_id, VAK_NAMEN.get(
        sheet_naam.lower() if sheet_naam else '',
        vak_id.replace('doelenset-bao-', '').replace('-', ' ').title()
    ))

    logger.info(
        f"XLSX geconverteerd: {bestandsnaam} → {vak_id} "
        f"({len(rijen)} rijen, {aantal_doelzinnen} doelzinnen, bronDatum: {bron_datum})"
    )

    return {
        'vak':              vak_id,
        'vakNaam':          vak_naam,
        'versie':           '2025-01',
        'bronBestand':      bestandsnaam,
        'bronDatum':        bron_datum,   # 'gewijzigd' timestamp uit Excel metadata
        'aantalRijen':      len(rijen),
        'aantalDoelzinnen': aantal_doelzinnen,
        'rijen':            rijen,
    }


def valideer_xlsx_bestand(bestandsnaam: str) -> list:
    """Snelle validatie van de bestandsnaam. Geeft lijst van foutmeldingen."""
    fouten = []
    if not bestandsnaam.lower().endswith('.xlsx'):
        fouten.append("Alleen .xlsx bestanden zijn toegestaan (geen .xls of .csv)")
    return fouten
